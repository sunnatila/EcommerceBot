import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib import messages
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path, reverse


class ExcelExportMixin:
    change_list_template = 'admin/excel_change_list.html'
    excel_filename = 'export'
    excel_headers = []
    excel_import_enabled = False  # Import kerak bo'lgan adminda True qiling

    def get_urls(self):
        app = self.model._meta.app_label
        model = self.model._meta.model_name
        urls = [
            path(
                'export-excel/',
                self.admin_site.admin_view(self.export_to_excel),
                name=f'{app}_{model}_export_excel',
            ),
        ]
        if self.excel_import_enabled:
            urls.append(
                path(
                    'import-excel/',
                    self.admin_site.admin_view(self.import_from_excel),
                    name=f'{app}_{model}_import_excel',
                )
            )
        return urls + super().get_urls()

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        app = self.model._meta.app_label
        model = self.model._meta.model_name
        extra_context['export_excel_url'] = reverse(f'admin:{app}_{model}_export_excel')
        if self.excel_import_enabled:
            extra_context['import_excel_url'] = reverse(f'admin:{app}_{model}_import_excel')
        return super().changelist_view(request, extra_context=extra_context)

    # ── EXPORT ──────────────────────────────────────────────────────────────

    def export_to_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.excel_filename[:31]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        thin = Side(style='thin', color='BBBBBB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, (label, _) in enumerate(self.excel_headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        queryset = self.get_queryset(request)
        for row, obj in enumerate(queryset, 2):
            for col, (_, field) in enumerate(self.excel_headers, 1):
                if callable(field):
                    value = field(obj)
                else:
                    raw = getattr(obj, field, '')
                    value = raw() if callable(raw) else raw
                cell = ws.cell(row=row, column=col, value=str(value) if value is not None else '')
                cell.border = border
                cell.alignment = left

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.excel_filename}.xlsx"'
        wb.save(response)
        return response

    # ── IMPORT ──────────────────────────────────────────────────────────────

    def import_excel_row(self, row):
        """Har bir qatorni qayta ishlash. Subclassda override qiling."""
        raise NotImplementedError

    def import_from_excel(self, request):
        app = self.model._meta.app_label
        model = self.model._meta.model_name

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Fayl tanlanmadi.")
                return _redirect_to_changelist(app, model)

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f"Excel faylni o'qib bo'lmadi: {e}")
                return _redirect_to_changelist(app, model)

            imported, errors = 0, []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if all(v is None for v in row):
                    continue
                try:
                    self.import_excel_row(list(row))
                    imported += 1
                except Exception as e:
                    errors.append(f"Qator {row_num}: {e}")

            if errors:
                for err in errors[:10]:
                    messages.warning(request, err)
            messages.success(request, f"{imported} ta qator muvaffaqiyatli import qilindi.")
            return _redirect_to_changelist(app, model)

        context = self.admin_site.each_context(request)
        context['title'] = f"Excel import — {self.model._meta.verbose_name_plural}"
        context['opts'] = self.model._meta
        context['excel_filename'] = self.excel_filename
        return TemplateResponse(request, 'admin/excel_import.html', context)


def _redirect_to_changelist(app, model):
    from django.shortcuts import redirect
    return redirect(reverse(f'admin:{app}_{model}_changelist'))
