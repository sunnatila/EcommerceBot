import os

from aiogram import types
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from handlers.users.start import AdminFilter
from loader import dp, db

FILE_PATH = "./data/start_bosganlar.xlsx"


@dp.message(AdminFilter(), lambda msg: msg.text == "📊 Start bosganlar")
async def show_starts(message: types.Message):
    users = await db.get_all_starts()
    if not users:
        await message.answer("Hali hech kim /start bosmagan.")
        return

    # data papka yo'q bo'lsa yaratamiz
    os.makedirs("data", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Start bosganlar"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["№", "Telegram ID", "Ism-familiya", "Username", "Sana"]
    widths = [6, 18, 25, 20, 22]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[chr(64 + col)].width = width
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, (tg_id, fullname, username, started_at) in enumerate(users, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=tg_id).border = thin_border
        ws.cell(row=row, column=3, value=fullname or "—").border = thin_border
        ws.cell(row=row, column=4, value=f"@{username}" if username else "—").border = thin_border
        ws.cell(row=row, column=5, value=str(started_at)).border = thin_border

    # Eski fayl bo'lsa ustiga yozadi, yo'q bo'lsa yangi yaratadi
    wb.save(FILE_PATH)

    # Faylni yuborish
    await message.answer_document(
        document=types.FSInputFile(path=FILE_PATH, filename=f"start_bosganlar.xlsx"),
        caption=f"📊 Jami: {len(users)} ta foydalanuvchi"
    )

